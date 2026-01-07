"""CSV and Excel export functionality."""
import csv
from pathlib import Path
from typing import List

from ..core.models import PriceTimeSeries, PriceSnapshot


class CSVExporter:
    """Export price data to CSV format."""
    
    @staticmethod
    def export_timeseries(
        timeseries: PriceTimeSeries,
        output_path: Path,
        include_metadata: bool = True
    ):
        """
        Export time series to CSV.
        
        Args:
            timeseries: PriceTimeSeries to export
            output_path: Output file path
            include_metadata: Include extraction metadata columns
        """
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            headers = [
                'Date',
                'Price',
                'Currency',
                'Type',
                'Tier',
                'Exact Match',
                'Days Offset',
            ]
            
            if include_metadata:
                headers.extend([
                    'Confidence',
                    'Method',
                    'Wayback URL',
                ])
            
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            
            for ps in timeseries.snapshots:
                if not ps.has_prices:
                    # Write row with no price to show gap
                    row = {
                        'Date': ps.snapshot.timestamp.strftime('%Y-%m-%d'),
                        'Price': 'N/A',
                        'Currency': '',
                        'Type': '',
                        'Tier': '',
                        'Exact Match': ps.snapshot.is_exact_match,
                        'Days Offset': ps.snapshot.distance_days,
                    }
                    if include_metadata:
                        row.update({
                            'Confidence': '',
                            'Method': '',
                            'Wayback URL': ps.snapshot.wayback_url,
                        })
                    writer.writerow(row)
                    continue
                
                for price in ps.prices:
                    row = {
                        'Date': ps.snapshot.timestamp.strftime('%Y-%m-%d'),
                        'Price': f'{price.value:.2f}',
                        'Currency': price.currency.value,
                        'Type': price.price_type.value,
                        'Tier': price.tier_name or '',
                        'Exact Match': ps.snapshot.is_exact_match,
                        'Days Offset': ps.snapshot.distance_days,
                    }
                    
                    if include_metadata:
                        row.update({
                            'Confidence': f'{price.confidence:.2f}',
                            'Method': price.extraction_method.value,
                            'Wayback URL': ps.snapshot.wayback_url,
                        })
                    
                    writer.writerow(row)


class ExcelExporter:
    """Export price data to Excel format."""
    
    @staticmethod
    def export_timeseries(
        timeseries: PriceTimeSeries,
        output_path: Path,
        include_charts: bool = True
    ):
        """
        Export time series to Excel with formatting.
        
        Args:
            timeseries: PriceTimeSeries to export
            output_path: Output file path
            include_charts: Add charts to workbook
        """
        try:
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Price History"
        
        # Headers
        headers = [
            'Date', 'Price', 'Currency', 'Type', 'Tier',
            'Exact Match', 'Days Offset', 'Confidence', 'Method'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 2
        for ps in timeseries.snapshots:
            if not ps.has_prices:
                ws.cell(row=row, column=1, value=ps.snapshot.timestamp.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value='N/A')
                row += 1
                continue
            
            for price in ps.prices:
                ws.cell(row=row, column=1, value=ps.snapshot.timestamp.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=price.value)
                ws.cell(row=row, column=3, value=price.currency.value)
                ws.cell(row=row, column=4, value=price.price_type.value)
                ws.cell(row=row, column=5, value=price.tier_name or '')
                ws.cell(row=row, column=6, value='Yes' if ps.snapshot.is_exact_match else 'No')
                ws.cell(row=row, column=7, value=ps.snapshot.distance_days)
                ws.cell(row=row, column=8, value=price.confidence)
                ws.cell(row=row, column=9, value=price.extraction_method.value)
                row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add chart if requested
        if include_charts and row > 2:
            chart = LineChart()
            chart.title = "Price Over Time"
            chart.style = 13
            chart.y_axis.title = 'Price'
            chart.x_axis.title = 'Date'
            
            data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
            dates = Reference(ws, min_col=1, min_row=2, max_row=row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(dates)
            
            ws.add_chart(chart, f'K2')
        
        wb.save(output_path)
